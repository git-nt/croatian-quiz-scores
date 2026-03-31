"""Load rank.xlsx from per-event sheets only. Medals and summaries are derived from placements."""

from __future__ import annotations

import json
import re
import statistics
import unicodedata
from dataclasses import dataclass, field
from functools import cached_property
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_BASE = Path(__file__).resolve().parent
EXCEL_PATH = _BASE / "rank.xlsx"

SKIP_SHEET_NAMES = frozenset({"results", "medals"})


# ── Series config loaded from series.json ──

@dataclass
class SeriesInfo:
    code: str
    name: str
    blurb: str
    month: int
    month_overrides: dict[int, int] = field(default_factory=dict)
    defunct: bool = False
    category: str = "live"  # "live" | "hks"
    location: str = ""

    def month_for_year(self, year: int) -> int:
        return self.month_overrides.get(year, self.month)


def _load_series() -> list[SeriesInfo]:
    with open(_BASE / "series.json", encoding="utf-8") as f:
        raw = json.load(f)
    out: list[SeriesInfo] = []
    for entry in raw:
        overrides = {int(k): v for k, v in entry.get("month_overrides", {}).items()}
        out.append(SeriesInfo(
            code=entry["code"],
            name=entry["name"],
            blurb=entry["blurb"],
            month=entry["month"],
            month_overrides=overrides,
            defunct=entry.get("defunct", False),
            category=entry.get("category", "live"),
            location=entry.get("location", ""),
        ))
    return out


SERIES: list[SeriesInfo] = _load_series()
SERIES_BY_CODE: dict[str, SeriesInfo] = {s.code: s for s in SERIES}


def norm_alnum(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _diacritic_score(s: str) -> int:
    """Count non-ASCII characters — higher means more diacritics, i.e. more canonical name."""
    return sum(1 for c in s if ord(c) > 127)


def parse_rank(val: Any) -> int | None:
    if val is None or val == "" or val == "-":
        return None
    if isinstance(val, str):
        v = val.strip()
        if v in ("-", "#NUM!", "#N/A", "#VALUE!"):
            return None
        try:
            return int(float(v.replace(",", ".")))
        except (ValueError, TypeError):
            return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            return int(val)
        except (ValueError, TypeError):
            return None
    return None


def parse_points(val: Any) -> float | None:
    """Numeric score for columns C/D (final+regular or regular-only)."""
    if val is None or val == "" or val == "-":
        return None
    if isinstance(val, str):
        v = val.strip()
        if v in ("-", "#NUM!", "#N/A", "#VALUE!"):
            return None
        try:
            return float(v.replace(",", "."))
        except (ValueError, TypeError):
            return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    return None


def player_match_key(name: str) -> str:
    """Normalize for cross-sheet identity (hyphens, spaces, diacritics)."""
    s = unicodedata.normalize("NFKD", str(name))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("\xa0", " ").strip().lower()
    s = re.sub(r"[-\u2010-\u2015]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def tournament_key_from_label(label: str) -> str:
    s = str(label).strip()
    if re.search(r"osijek|^OO\s*\d", s, re.I):
        m = re.search(r"(\d{4})", s)
        return f"oo{m.group(1)}" if m else norm_alnum(s)
    if re.search(r"ZIMKO", s, re.I):
        m = re.search(r"(\d{4})", s)
        return f"zimko{m.group(1)}" if m else norm_alnum(s)
    if re.search(r"OVaKO|OVAKO", s, re.I) or norm_alnum(s).startswith("ovako"):
        m = re.search(r"(\d{4})", s)
        return f"ovako{m.group(1)}" if m else norm_alnum(s)
    if re.search(r"MMM", s, re.I):
        m = re.search(r"(\d{4})", s)
        return f"mmm{m.group(1)}" if m else norm_alnum(s)
    if re.search(r"SOVA", s, re.I):
        m = re.search(r"(\d{4})", s)
        return f"sova{m.group(1)}" if m else norm_alnum(s)
    if re.search(r"Croatian\s*Open|CRO\s*Open", s, re.I):
        m = re.search(r"(\d{4})", s)
        return f"cro{m.group(1)}" if m else norm_alnum(s)
    return norm_alnum(s)


_SERIES_CODES_LONGEST_FIRST = sorted(
    [s.code for s in SERIES], key=len, reverse=True
)


def series_code_from_key(key: str) -> str:
    for code in _SERIES_CODES_LONGEST_FIRST:
        if key.startswith(code):
            return code
    return key[:4] if len(key) >= 4 else key


def blurb_for_key(key: str) -> str | None:
    info = SERIES_BY_CODE.get(series_code_from_key(key))
    return info.blurb if info else None


def tournament_sort_key(key: str) -> tuple[int, int]:
    """(year, month) for chronological sorting, derived from series.json."""
    code = series_code_from_key(key)
    m = re.search(r"(\d{4})", key)
    year = int(m.group(1)) if m else 0
    info = SERIES_BY_CODE.get(code)
    month = info.month_for_year(year) if info else 6
    return (year, month)


@dataclass
class TournamentColumn:
    key: str
    label: str

    @property
    def year(self) -> str | None:
        m = re.search(r"(\d{4})", self.label)
        return m.group(1) if m else None


@dataclass
class SheetStandingRow:
    """One row from an event sheet: rank, name, total (final+qualifying), regular (qualifying-only for finalists)."""

    rank: int | None
    name: str
    total: Any  # combined score, display as-is
    regular: Any | None  # qualifying-phase score when finalist; else None
    is_finalist: bool


@dataclass
class PlayerEventDetail:
    """Single event for a player: placement and scores from the sheet."""

    rank: int
    total: Any
    regular: Any | None
    is_finalist: bool


@dataclass
class PlayerResult:
    name: str
    ranks: dict[str, int] = field(default_factory=dict)
    event_details: dict[str, PlayerEventDetail] = field(default_factory=dict)

    def medal_counts(self) -> tuple[int, int, int]:
        g = s = b = 0
        for r in self.ranks.values():
            if r == 1:
                g += 1
            elif r == 2:
                s += 1
            elif r == 3:
                b += 1
        return g, s, b

    @property
    def best_placement(self) -> int | None:
        vals = [r for r in self.ranks.values() if r is not None]
        return min(vals) if vals else None

    @property
    def median_placement(self) -> float | None:
        vals = sorted([r for r in self.ranks.values() if r is not None])
        if not vals:
            return None
        return float(statistics.median(vals))


@dataclass
class RankingData:
    tournaments: list[TournamentColumn]
    players: list[PlayerResult]
    sheet_by_key: dict[str, str]

    @cached_property
    def tournament_labels(self) -> dict[str, str]:
        return {t.key: t.label for t in self.tournaments}

    def standings_rows(self, key: str) -> list[SheetStandingRow]:
        """Full standings: C = total (final+regular for finalists), D = regular phase (finalists only)."""
        sheet = self.sheet_by_key.get(key)
        if not sheet:
            return []
        wb = load_workbook(EXCEL_PATH, data_only=True)
        try:
            ws = wb[sheet]
            out: list[SheetStandingRow] = []
            for rank, name, c_raw, d_raw in _iter_sheet_placement_rows(ws):
                c_pts = parse_points(c_raw)
                d_pts = parse_points(d_raw)
                is_finalist = c_pts is not None and d_pts is not None
                out.append(
                    SheetStandingRow(
                        rank=rank,
                        name=name,
                        total=c_raw,
                        regular=d_raw if is_finalist else None,
                        is_finalist=is_finalist,
                    )
                )
            return out
        finally:
            wb.close()


def _sheet_has_results(ws: Any) -> bool:
    """At least one data row with a name in column B (row 1 is first placement, no header row)."""
    for r in range(1, min(ws.max_row + 1, 500)):
        name = ws.cell(r, 2).value
        if name is not None and str(name).strip() != "":
            return True
    return False


def _iter_sheet_placement_rows(ws: Any):
    """Yield (rank, name, c_raw, d_raw) per data row.

    - Rank: column A when present; otherwise 1,2,3… by row order (e.g. SOVA 2017 has no ranks in A).
    - Points: columns C/D when present; may be missing (e.g. SOVA 2024 top 7 names only).
    """
    implicit = 0
    for r in range(1, ws.max_row + 1):
        raw_name = ws.cell(r, 2).value
        if raw_name is None or str(raw_name).strip() == "":
            continue
        implicit += 1
        rank = parse_rank(ws.cell(r, 1).value)
        if rank is None:
            rank = implicit
        name = str(raw_name).replace("\xa0", " ").strip()
        c_raw = ws.cell(r, 3).value if ws.max_column >= 3 else None
        d_raw = ws.cell(r, 4).value if ws.max_column >= 4 else None
        yield rank, name, c_raw, d_raw


def load_ranking_data(path: Path | None = None) -> RankingData:
    path = path or EXCEL_PATH
    wb = load_workbook(path, data_only=True)
    try:
        tournaments: list[TournamentColumn] = []
        sheet_by_key: dict[str, str] = {}

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in SKIP_SHEET_NAMES:
                continue
            ws = wb[sheet_name]
            if not _sheet_has_results(ws):
                continue
            key = tournament_key_from_label(sheet_name)
            if key in sheet_by_key:
                # Prefer first sheet; skip duplicate key (same event twice)
                continue
            sheet_by_key[key] = sheet_name
            tournaments.append(TournamentColumn(key=key, label=sheet_name))

        tournaments.sort(key=lambda t: tournament_sort_key(t.key))

        merged: dict[str, dict[str, Any]] = {}

        for t in tournaments:
            ws = wb[t.label]
            for rank, display, c_raw, d_raw in _iter_sheet_placement_rows(ws):
                c_pts = parse_points(c_raw)
                d_pts = parse_points(d_raw)
                is_finalist = c_pts is not None and d_pts is not None
                mk = player_match_key(display)
                if mk not in merged:
                    merged[mk] = {"name": display, "ranks": {}, "event_details": {}}
                else:
                    existing = merged[mk]["name"]
                    if _diacritic_score(display) > _diacritic_score(existing):
                        merged[mk]["name"] = display
                    elif _diacritic_score(display) == _diacritic_score(existing) and len(display) > len(existing):
                        merged[mk]["name"] = display
                merged[mk]["ranks"][t.key] = rank
                merged[mk]["event_details"][t.key] = PlayerEventDetail(
                    rank=rank,
                    total=c_raw,
                    regular=d_raw if is_finalist else None,
                    is_finalist=is_finalist,
                )

        players = [
            PlayerResult(
                name=entry["name"],
                ranks=entry["ranks"],
                event_details=entry.get("event_details", {}),
            )
            for entry in merged.values()
        ]
        players.sort(key=lambda p: p.name.lower())

        return RankingData(tournaments=tournaments, players=players, sheet_by_key=sheet_by_key)
    finally:
        wb.close()


def aggregate_medals(
    data: RankingData,
    filter_codes: set[str] | None = None,
) -> list[tuple[str, int, int, int, int]]:
    """Medalists only: placement 1–3 on any event sheet (G/S/B). Sorted by Σ, then G.

    If filter_codes is given, only count events whose series code is in that set.
    Excludes players with no podiums — derived from `PlayerResult.ranks`, not any summary sheet.
    """
    rows = []
    for p in data.players:
        if filter_codes is None:
            g, s, b = p.medal_counts()
        else:
            g = s = b = 0
            for tkey, rank in p.ranks.items():
                if series_code_from_key(tkey) not in filter_codes:
                    continue
                if rank == 1:
                    g += 1
                elif rank == 2:
                    s += 1
                elif rank == 3:
                    b += 1
        if g + s + b == 0:
            continue
        rows.append((p.name, g, s, b, g + s + b))
    rows.sort(key=lambda x: (-x[1], -x[2], -x[3], -x[4], x[0].lower()))
    return rows


def medalists_by_event(
    data: RankingData,
) -> list[tuple[TournamentColumn, list[tuple[str, int]]]]:
    """Each tournament with list of (player_name, rank) for rank <= 3."""
    out: list[tuple[TournamentColumn, list[tuple[str, int]]]] = []
    for t in data.tournaments:
        podium: list[tuple[str, int]] = []
        for p in data.players:
            r = p.ranks.get(t.key)
            if r is not None and 1 <= r <= 3:
                podium.append((p.name, r))
        podium.sort(key=lambda x: (x[1], x[0].lower()))
        out.append((t, podium))
    return out
