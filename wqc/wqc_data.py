"""Load World Quiz Championship standings from wqc_history/wqc_scores.xlsx (decoupled from Croatian site)."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

WQC_ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = WQC_ROOT.parent / "wqc_history" / "wqc_scores.xlsx"


def norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = unicodedata.normalize("NFKD", str(h))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]+", "", s.lower().strip())
    return s


# (id, display name, normalized header aliases)
GENRES: list[tuple[str, str, frozenset[str]]] = [
    ("cul", "Culture", frozenset({
        "cul", "cult", "culture", "clt", "articulture", "artculture",
    })),
    ("ent", "Entertainment", frozenset({
        "ent", "entertainment", "enter",
    })),
    ("his", "History", frozenset({
        "his", "hst", "hist", "history", "histor", "civilisation", "civilization",
    })),
    ("med", "Media", frozenset({
        "med", "media",
    })),
    ("lif", "Lifestyle", frozenset({
        "lif", "lfs", "lifest", "lifestyle", "life",
    })),
    ("sci", "Science", frozenset({
        "sci", "scien", "science", "scie",
    })),
    ("spo", "Sport & Games", frozenset({
        "spo", "spt", "sport", "sports", "sportleisure", "sportgames",
    })),
    ("wor", "World", frozenset({
        "wor", "wld", "world", "physicalworld",
    })),
]

GENRE_IDS = [g[0] for g in GENRES]
GENRE_LABEL = {g[0]: g[1] for g in GENRES}


def player_slug(name: str) -> str:
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "player"


def parse_num(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, str):
        v = val.strip()
        if v in ("-", "#N/A", "#VALUE!", "#NUM!"):
            return None
        try:
            return float(v.replace(",", "."))
        except (ValueError, TypeError):
            return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    return None


def parse_int_rank(val: Any) -> int | None:
    n = parse_num(val)
    if n is None:
        return None
    return int(n)


@dataclass
class StandingRow:
    name: str
    rank: int
    score: float | None
    genres: dict[str, float | None] = field(default_factory=dict)
    genre_ranks: dict[str, int | None] = field(default_factory=dict)


@dataclass
class YearSheet:
    year: int
    rows: list[StandingRow]
    genres_present: frozenset[str]


@dataclass
class PlayerCareer:
    name: str
    by_year: dict[int, StandingRow]


def _map_headers(headers: list[Any]) -> tuple[int | None, int | None, int | None, dict[str, int]]:
    """Rank col, Player col, SCORE col, genre_id -> col index."""
    norms = [norm_header(h) for h in headers]
    rank_i = next((i for i, n in enumerate(norms) if n == "rank"), None)
    player_i = next((i for i, n in enumerate(norms) if n == "player"), None)
    score_i = None
    for prefer in ("score", "total"):
        for i, n in enumerate(norms):
            if n == prefer:
                score_i = i
                break
        if score_i is not None:
            break

    genre_col: dict[str, int] = {}
    used: set[int] = set()
    if rank_i is not None:
        used.add(rank_i)
    if player_i is not None:
        used.add(player_i)
    if score_i is not None:
        used.add(score_i)

    for gid, _label, aliases in GENRES:
        for i, n in enumerate(norms):
            if i in used:
                continue
            if n in aliases:
                genre_col[gid] = i
                used.add(i)
                break

    return rank_i, player_i, score_i, genre_col


def _sheet_year(name: str) -> int | None:
    m = re.search(r"(\d{4})", name)
    return int(m.group(1)) if m else None


def load_wqc_data(xlsx_path: Path | None = None, *, progress: bool = True) -> tuple[list[int], dict[int, YearSheet], dict[str, PlayerCareer], dict[int, frozenset[str]]]:
    path = xlsx_path or DEFAULT_XLSX
    if not path.exists():
        raise FileNotFoundError(f"WQC Excel not found: {path}")

    def _p(msg: str) -> None:
        if progress:
            print(msg, flush=True)

    _p(f"WQC: opening {path} …")
    wb = load_workbook(path, data_only=True)
    year_sheets: dict[int, YearSheet] = {}

    wqc_names = [n for n in wb.sheetnames if "WQC" in n.upper() and _sheet_year(n) is not None]
    _p(f"WQC: {len(wqc_names)} championship sheets to parse …")

    try:
        for snum, sheet_name in enumerate(wqc_names, start=1):
            year = _sheet_year(sheet_name)
            if year is None:
                continue
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                _p(f"WQC:   [{snum}/{len(wqc_names)}] {sheet_name} — skipped (empty)")
                continue
            rank_i, player_i, score_i, genre_col = _map_headers(list(header))
            if rank_i is None or player_i is None:
                _p(f"WQC:   [{snum}/{len(wqc_names)}] {sheet_name} — skipped (no Rank/Player columns)")
                continue

            raw_rows: list[StandingRow] = []
            for row in rows_iter:
                if not row or row[player_i] is None:
                    continue
                name = str(row[player_i]).strip()
                if not name or name.lower() == "player":
                    continue
                rk = parse_int_rank(row[rank_i])
                if rk is None:
                    continue
                sc = None
                if score_i is not None and score_i < len(row):
                    sc = parse_num(row[score_i])
                gvals: dict[str, float | None] = {}
                for gid, ci in genre_col.items():
                    if ci < len(row):
                        gvals[gid] = parse_num(row[ci])
                    else:
                        gvals[gid] = None
                raw_rows.append(StandingRow(name=name, rank=rk, score=sc, genres=gvals))

            present = frozenset(genre_col.keys())
            for gid in GENRE_IDS:
                scores: list[tuple[str, float]] = []
                for r in raw_rows:
                    v = r.genres.get(gid)
                    if v is not None:
                        scores.append((r.name, v))
                scores.sort(key=lambda x: -x[1])
                # Competition rank: 1 + count strictly higher
                rank_map: dict[str, int] = {}
                for i, (nm, val) in enumerate(scores):
                    higher = sum(1 for _nm, v in scores if v > val)
                    rank_map[nm] = higher + 1
                for r in raw_rows:
                    r.genre_ranks[gid] = rank_map.get(r.name)

            year_sheets[year] = YearSheet(year=year, rows=raw_rows, genres_present=present)
            _p(f"WQC:   [{snum}/{len(wqc_names)}] {sheet_name} — {len(raw_rows)} players")
    finally:
        wb.close()

    years_sorted = sorted(year_sheets.keys())
    players: dict[str, PlayerCareer] = {}

    _p("WQC: merging player careers …")
    for y in years_sorted:
        ys = year_sheets[y]
        for r in ys.rows:
            if r.name not in players:
                players[r.name] = PlayerCareer(name=r.name, by_year={})
            players[r.name].by_year[y] = r

    genres_by_year = {y: year_sheets[y].genres_present for y in years_sorted}
    _p(f"WQC: ready — {len(years_sorted)} years, {len(players)} unique names.")
    return years_sorted, year_sheets, players, genres_by_year


def medal_counts_for(rows: list[tuple[str, int]]) -> tuple[int, int, int]:
    g = s = b = 0
    for _name, r in rows:
        if r == 1:
            g += 1
        elif r == 2:
            s += 1
        elif r == 3:
            b += 1
    return g, s, b


def aggregate_medals(
    year_sheets: dict[int, YearSheet],
    *,
    kind: str = "overall",
) -> list[tuple[str, int, int, int, int]]:
    """kind: 'overall' or any GENRE id."""
    per_player: dict[str, list[int]] = {}

    for y, ys in year_sheets.items():
        for r in ys.rows:
            if kind == "overall":
                rk = r.rank
            else:
                if kind not in ys.genres_present:
                    continue
                gr = r.genre_ranks.get(kind)
                if gr is None:
                    continue
                rk = gr
            if rk <= 3:
                per_player.setdefault(r.name, []).append(rk)

    out: list[tuple[str, int, int, int, int]] = []
    for name, ranks in per_player.items():
        g = sum(1 for x in ranks if x == 1)
        s = sum(1 for x in ranks if x == 2)
        b = sum(1 for x in ranks if x == 3)
        out.append((name, g, s, b, g + s + b))
    out.sort(key=lambda x: (-x[1], -x[2], -x[3], x[0].lower()))
    return out


def podium_by_year_overall(year_sheets: dict[int, YearSheet]) -> list[tuple[int, list[tuple[str, int]]]]:
    items: list[tuple[int, list[tuple[str, int]]]] = []
    for y in sorted(year_sheets.keys(), reverse=True):
        rows = [(r.name, r.rank) for r in year_sheets[y].rows if r.rank <= 3]
        rows.sort(key=lambda t: t[1])
        items.append((y, rows))
    return items


def podium_by_year_genre(
    year_sheets: dict[int, YearSheet],
    genre: str,
) -> list[tuple[int, list[tuple[str, int]]]]:
    """Top 3 in a genre column for each year (year descending)."""
    items: list[tuple[int, list[tuple[str, int]]]] = []
    for y in sorted(year_sheets.keys(), reverse=True):
        ys = year_sheets[y]
        if genre not in ys.genres_present:
            items.append((y, []))
            continue
        rows: list[tuple[str, int]] = []
        for r in ys.rows:
            gr = r.genre_ranks.get(genre)
            if gr is not None and gr <= 3:
                rows.append((r.name, gr))
        rows.sort(key=lambda t: t[1])
        items.append((y, rows))
    return items


def player_wqc_podium_rows(career: PlayerCareer, genre: str | None) -> list[dict[str, int]]:
    """Each entry: year, rank — overall if genre is None, else genre rank.

    Sorted: all golds (1st) ascending by year, then silvers (2nd) ascending by year,
    then bronzes (3rd) ascending by year.
    """
    out: list[dict[str, int]] = []
    for y in career.by_year.keys():
        r = career.by_year[y]
        if genre is None:
            if r.rank <= 3:
                out.append({"year": y, "rank": r.rank})
        else:
            gr = r.genre_ranks.get(genre)
            if gr is not None and gr <= 3:
                out.append({"year": y, "rank": gr})
    out.sort(key=lambda d: (d["rank"], d["year"]))
    return out


def timeline_css(rank: int | None, *, has_genre_column: bool, participated: bool) -> str:
    if not participated:
        return "tl-absent"
    if not has_genre_column:
        return "tl-no-event"
    if rank is None:
        return "tl-no-event"
    if rank == 1:
        return "tl-gold"
    if rank == 2:
        return "tl-silver"
    if rank == 3:
        return "tl-bronze"
    if rank <= 10:
        return "tl-final"
    return "tl-played"


def build_player_timeline(
    career: PlayerCareer,
    all_years: list[int],
    year_sheets: dict[int, YearSheet],
) -> tuple[list[int], list[dict], list[dict]]:
    """Returns years, overall row spec, list of genre row specs."""
    rows_overall: list[dict] = []
    for y in all_years:
        ys = year_sheets.get(y)
        if ys is None:
            rows_overall.append({"year": y, "state": "no_sheet"})
            continue
        r = career.by_year.get(y)
        if r is None:
            rows_overall.append({"year": y, "state": "absent"})
        else:
            rows_overall.append({
                "year": y,
                "state": "played",
                "rank": r.rank,
                "css": timeline_css(r.rank, has_genre_column=True, participated=True),
            })

    genre_rows: list[dict] = []
    for gid, label, _aliases in GENRES:
        cells: list[dict] = []
        for y in all_years:
            ys = year_sheets.get(y)
            if ys is None:
                cells.append({"year": y, "state": "no_sheet"})
                continue
            has_col = gid in ys.genres_present
            r = career.by_year.get(y)
            if r is None:
                cells.append({"year": y, "state": "absent"})
                continue
            if not has_col:
                cells.append({"year": y, "state": "na", "css": "tl-no-event"})
                continue
            gr = r.genre_ranks.get(gid)
            if gr is None:
                cells.append({"year": y, "state": "na", "css": "tl-no-event"})
            else:
                cells.append({
                    "year": y,
                    "state": "played",
                    "rank": gr,
                    "css": timeline_css(gr, has_genre_column=True, participated=True),
                })
        genre_rows.append({"id": gid, "label": label, "cells": cells})

    return all_years, {"label": "Overall", "cells": rows_overall}, genre_rows
